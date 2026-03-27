"""Analyse the multi-agent pipeline stage by stage.

Outputs
-------
1. em_ba_by_stage.png      — EM & BA across pipeline stages (Figure 4)
2. stage_transitions_*.csv — every instance with per-stage predictions & transitions
3. Console: cross-model comparison table and label-change statistics (Table 6)

Usage:
  python visualize_pipeline_flow.py
  python visualize_pipeline_flow.py --models GPT,Claude
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import typer

app = typer.Typer()

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_PATH = PROJECT_ROOT / "data" / "swe_gym_all.csv"
SUB_AGENT_DIR = PROJECT_ROOT / "data" / "sub_agent_outputs"
OUTPUT_DIR = PROJECT_ROOT / "data" / "outputs" / "checker"

MODEL_DIRS: dict[str, str] = {
    "GPT": "openai_gpt-5.3-codex",
    "Gemini": "openai_gemini-3.1-pro-preview",
    "Qwen": "openai_qwen3.5-plus",
    "Claude": "openai_claude-sonnet-4-6",
}
MODEL_SLUG: dict[str, str] = {
    "GPT": "gpt",
    "Gemini": "gemini",
    "Qwen": "qwen",
    "Claude": "claude",
}
AGENT_FILES = [
    "issue_analyzer.txt",
    "pr_scope_analyzer.txt",
    "pr_connection_analyzer.txt",
    "coordinator.txt",
]
ALL_LABELS = ["SC", "FP", "DP", "IS", "UL", "Others", "No Misalignment"]
STAGE_KEYS = ("tier1", "coord", "cv")
STAGE_NAMES = ["Tier-1", "Coordinator", "Code Validator"]


# ── Data structures ───────────────────────────────────────────────────────


@dataclass
class Instance:
    iid: str
    human: set[str]
    tier1: set[str]
    coord: set[str]
    cv: set[str]
    dim_ok: list[bool]  # [IA, PSA, PCA, coord_em, cv_em]


@dataclass
class Transition:
    cc: int = 0
    ci: int = 0
    ic: int = 0
    ii: int = 0

    @property
    def net(self) -> int:
        return self.ic - self.ci


@dataclass
class LabelChange:
    changed: int = 0
    fixed: int = 0
    broken: int = 0

    @property
    def fix_rate(self) -> float:
        return self.fixed / self.changed if self.changed else 0.0

    @property
    def break_rate(self) -> float:
        return self.broken / self.changed if self.changed else 0.0


@dataclass
class Metrics:
    n: int = 0
    em: list[float] = field(default_factory=lambda: [0.0, 0.0, 0.0])
    ba: list[float] = field(default_factory=lambda: [0.0, 0.0, 0.0])
    per_label_f1: dict[str, list[float]] = field(default_factory=dict)
    t1_coord: Transition = field(default_factory=Transition)
    coord_cv: Transition = field(default_factory=Transition)
    label_change: dict[str, LabelChange] = field(default_factory=dict)


# ── Data loading ──────────────────────────────────────────────────────────


def _parse_label(raw) -> set[str]:
    if not raw or str(raw).strip() == "":
        return {"No Misalignment"}
    return {re.sub(r"-\d+$", "", p.strip()) for p in str(raw).split(",")}


def load_human_labels(path: Path = DATA_PATH) -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    if path.suffix.lower() in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        id_col, label_col = headers.index("instance_id"), headers.index("human_label")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[id_col]:
                result[row[id_col]] = _parse_label(row[label_col])
        wb.close()
    else:
        for row in csv.DictReader(path.open(encoding="utf-8")):
            if iid := row.get("instance_id", "").strip():
                result[iid] = _parse_label(row.get("human_label", ""))
    return result


def _tag(text: str, tag: str) -> str:
    m = re.search(rf"<{tag}>(.*?)</{tag}>", text, re.DOTALL | re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _load_jsonl(model_key: str) -> dict[str, dict]:
    path = OUTPUT_DIR / f"misalign_outputs_{MODEL_SLUG[model_key]}_all.jsonl"
    if not path.exists():
        return {}
    return {
        rec["instance_id"]: rec
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
        for rec in [json.loads(line)]
        if rec.get("status") != "skipped" and rec.get("instance_id")
    }


def _jsonl_labels(rec: dict) -> set[str]:
    cls = rec.get("classifications", [])
    return {c["label"] for c in cls} if cls else {"No Misalignment"}


_EMPTY: set[str] = set()


def _parse_instance(iid: str, human: set[str], inst_dir: Path | None, rec: dict | None) -> Instance:
    tier1 = _EMPTY
    coord = _EMPTY
    dim_ok = [False, False, False, False, False]
    is_pos = sc_pos = ul_pos = dp_pos = fp_pos = False

    if inst_dir and inst_dir.is_dir():
        ia_path = inst_dir / "issue_analyzer.txt"
        psa_path = inst_dir / "pr_scope_analyzer.txt"
        pca_path = inst_dir / "pr_connection_analyzer.txt"
        coord_path = inst_dir / "coordinator.txt"

        if ia_path.exists() and psa_path.exists() and pca_path.exists():
            try:
                ia = ia_path.read_text(errors="replace")
                psa = psa_path.read_text(errors="replace")
                pca = pca_path.read_text(errors="replace")
                is_pos = _tag(ia, "judgment").upper() == "IS"
                sc_pos = _tag(psa, "sc_judgment").upper() == "SC"
                ul_pos = _tag(psa, "ul_judgment").upper() == "UL"
                dp_pos = _tag(pca, "dp_judgment").upper() == "DP"
                fp_pos = _tag(pca, "fp_judgment").upper() == "FP"
                tier1 = {lbl for flag, lbl in [(is_pos, "IS"), (sc_pos, "SC"), (ul_pos, "UL"), (dp_pos, "DP"), (fp_pos, "FP")] if flag}
                tier1 = tier1 or {"No Misalignment"}
            except Exception:
                pass

        if coord_path.exists():
            try:
                coord_txt = coord_path.read_text(errors="replace")
                labels: set[str] = set()
                for m in re.finditer(r"<classification>.*?<label>(.*?)</label>.*?</classification>", coord_txt, re.DOTALL):
                    labels.add(m.group(1).strip())
                coord = labels or {"No Misalignment"}
            except Exception:
                pass

    cv = _jsonl_labels(rec) if rec else _EMPTY

    return Instance(
        iid=iid, human=human,
        tier1=tier1, coord=coord, cv=cv,
        dim_ok=[
            is_pos == ("IS" in human),
            (sc_pos == ("SC" in human)) and (ul_pos == ("UL" in human)),
            (dp_pos == ("DP" in human)) and (fp_pos == ("FP" in human)),
            coord == human,
            cv == human,
        ],
    )


def load_model_data(model_key: str, human: dict[str, set[str]]) -> list[Instance]:
    jsonl = _load_jsonl(model_key)
    model_dir = SUB_AGENT_DIR / MODEL_DIRS[model_key]
    return [
        _parse_instance(iid, h, model_dir / iid, jsonl.get(iid))
        for iid, h in sorted(human.items())
    ]


# ── Metrics ───────────────────────────────────────────────────────────────


def _label_f1(instances: list[Instance], stage: str, label: str) -> float:
    tp = fp = fn = 0
    for inst in instances:
        pred = getattr(inst, stage)
        if label in inst.human and label in pred:
            tp += 1
        elif label not in inst.human and label in pred:
            fp += 1
        elif label in inst.human and label not in pred:
            fn += 1
    prec = tp / (tp + fp) if (tp + fp) else 0.0
    rec = tp / (tp + fn) if (tp + fn) else 0.0
    return 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0


def _make_transition(instances: list[Instance], a: str, b: str) -> Transition:
    t = Transition()
    for inst in instances:
        a_ok = getattr(inst, a) == inst.human
        b_ok = getattr(inst, b) == inst.human
        if a_ok and b_ok:       t.cc += 1
        elif a_ok and not b_ok: t.ci += 1
        elif not a_ok and b_ok: t.ic += 1
        else:                   t.ii += 1
    return t


def _make_label_change(instances: list[Instance], a: str, b: str) -> LabelChange:
    stats = LabelChange()
    for inst in instances:
        prev = getattr(inst, a)
        curr = getattr(inst, b)
        for label in ALL_LABELS:
            if (label in prev) == (label in curr):
                continue
            stats.changed += 1
            if (label in curr) == (label in inst.human):
                stats.fixed += 1
            else:
                stats.broken += 1
    return stats


def compute_metrics(instances: list[Instance]) -> Metrics:
    n = len(instances)
    if n == 0:
        return Metrics()

    em, ba = [], []
    for stage in STAGE_KEYS:
        em.append(sum(1 for i in instances if getattr(i, stage) == i.human) / n)
        ba.append(sum(
            1 for i in instances
            if (getattr(i, stage) == {"No Misalignment"}) == (i.human == {"No Misalignment"})
        ) / n)

    per_label: dict[str, list[float]] = {}
    for label in ALL_LABELS:
        per_label[label] = [_label_f1(instances, s, label) for s in STAGE_KEYS]

    return Metrics(
        n=n, em=em, ba=ba, per_label_f1=per_label,
        t1_coord=_make_transition(instances, "tier1", "coord"),
        coord_cv=_make_transition(instances, "coord", "cv"),
        label_change={
            "t1_coord": _make_label_change(instances, "tier1", "coord"),
            "coord_cv": _make_label_change(instances, "coord", "cv"),
        },
    )


# ── CSV output ────────────────────────────────────────────────────────────


def _fmt(labels: set[str]) -> str:
    order = {l: i for i, l in enumerate(ALL_LABELS)}
    return ",".join(sorted(labels, key=lambda x: order.get(x, 99)))


def _ttag(a: set[str], b: set[str], human: set[str]) -> str:
    a_ok, b_ok = a == human, b == human
    if a_ok and b_ok:       return "CC"
    if a_ok and not b_ok:   return "CI"
    if not a_ok and b_ok:   return "IC"
    return "II"


def write_csv(instances: list[Instance], model: str, out_dir: Path) -> None:
    path = out_dir / f"stage_transitions_{model}.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "instance_id", "human_label",
            "tier1_pred", "tier1_em",
            "coord_pred", "coord_em", "t1_coord",
            "cv_pred", "cv_em", "coord_cv",
            "ia_ok", "psa_ok", "pca_ok",
        ])
        for inst in sorted(instances, key=lambda x: x.iid):
            w.writerow([
                inst.iid, _fmt(inst.human),
                _fmt(inst.tier1), int(inst.tier1 == inst.human),
                _fmt(inst.coord), int(inst.coord == inst.human),
                _ttag(inst.tier1, inst.coord, inst.human),
                _fmt(inst.cv), int(inst.cv == inst.human),
                _ttag(inst.coord, inst.cv, inst.human),
                int(inst.dim_ok[0]), int(inst.dim_ok[1]), int(inst.dim_ok[2]),
            ])
    print(f"  CSV → {path}")


# ── Console output ────────────────────────────────────────────────────────


def print_model_report(model: str, m: Metrics) -> None:
    print(f"\n{'='*78}")
    print(f"  {model}  (n={m.n})")
    print(f"{'='*78}")

    print(f"\n  {'Stage':<18} {'EM':>7} {'ΔEM':>8} {'BA':>7} {'ΔBA':>8}")
    print(f"  {'-'*18} {'-'*7} {'-'*8} {'-'*7} {'-'*8}")
    for si, name in enumerate(STAGE_NAMES):
        d_em = f"{(m.em[si]-m.em[si-1])*100:+.1f}%" if si > 0 else ""
        d_ba = f"{(m.ba[si]-m.ba[si-1])*100:+.1f}%" if si > 0 else ""
        print(f"  {name:<18} {m.em[si]*100:>6.1f}% {d_em:>8} {m.ba[si]*100:>6.1f}% {d_ba:>8}")

    print(f"\n  Transition matrices (EM):")
    for label, t in [("Tier1 → Coord", m.t1_coord), ("Coord → CV", m.coord_cv)]:
        print(f"    {label}:  CC={t.cc}  CI={t.ci}  IC={t.ic}  II={t.ii}  net={t.net:+d}")

    print(f"\n  Label-change quality (only modified label slots):")
    for label, key in [("Tier1 → Coord", "t1_coord"), ("Coord → CV", "coord_cv")]:
        lc = m.label_change[key]
        print(
            f"    {label}:  changed={lc.changed}  fixed={lc.fixed}  broken={lc.broken}  "
            f"P(fix|change)={lc.fix_rate*100:.1f}%  P(break|change)={lc.break_rate*100:.1f}%"
        )

    active = [l for l in ALL_LABELS
              if any(f > 0 for f in m.per_label_f1.get(l, [0, 0, 0])) or l == "No Misalignment"]
    if active:
        print(f"\n  {'Label':<20} {'F1_T1':>6} {'F1_Co':>6} {'F1_CV':>6}  {'Δ(T1→Co)':>9} {'Δ(Co→CV)':>9}")
        print(f"  {'-'*20} {'-'*6} {'-'*6} {'-'*6}  {'-'*9} {'-'*9}")
        for label in active:
            f = m.per_label_f1[label]
            d1 = f"{(f[1]-f[0])*100:+.1f}%" if f[0] > 0 or f[1] > 0 else ""
            d2 = f"{(f[2]-f[1])*100:+.1f}%" if f[1] > 0 or f[2] > 0 else ""
            print(f"  {label:<20} {f[0]:>6.3f} {f[1]:>6.3f} {f[2]:>6.3f}  {d1:>9} {d2:>9}")


def print_cross_model_table(all_metrics: list[tuple[str, Metrics]]) -> None:
    print(f"\n{'='*110}")
    print("  CROSS-MODEL COMPARISON")
    print(f"{'='*110}")
    print(f"  {'Model':<10} {'N':>5}  "
          f"{'EM_T1':>6} {'EM_Co':>6} {'EM_CV':>6}  "
          f"{'BA_T1':>6} {'BA_Co':>6} {'BA_CV':>6}  "
          f"{'ΔEM(T→C)':>9} {'ΔEM(C→V)':>9} {'ΔBA(T→C)':>9} {'ΔBA(C→V)':>9}")
    print(f"  {'-'*10} {'-'*5}  {'-'*6} {'-'*6} {'-'*6}  {'-'*6} {'-'*6} {'-'*6}  {'-'*9} {'-'*9} {'-'*9} {'-'*9}")
    for name, m in all_metrics:
        print(
            f"  {name:<10} {m.n:>5}  "
            f"{m.em[0]*100:>5.1f}% {m.em[1]*100:>5.1f}% {m.em[2]*100:>5.1f}%  "
            f"{m.ba[0]*100:>5.1f}% {m.ba[1]*100:>5.1f}% {m.ba[2]*100:>5.1f}%  "
            f"{(m.em[1]-m.em[0])*100:>+8.1f}% {(m.em[2]-m.em[1])*100:>+8.1f}% "
            f"{(m.ba[1]-m.ba[0])*100:>+8.1f}% {(m.ba[2]-m.ba[1])*100:>+8.1f}%"
        )

    print(f"\n{'='*122}")
    print("  CROSS-MODEL LABEL-CHANGE QUALITY")
    print(f"{'='*122}")
    print(
        f"  {'Model':<10} "
        f"{'Chg T→C':>8} {'Fix% T→C':>10} {'Break% T→C':>12}  "
        f"{'Chg C→V':>8} {'Fix% C→V':>10} {'Break% C→V':>12}"
    )
    print(
        f"  {'-'*10} {'-'*8} {'-'*10} {'-'*12}  {'-'*8} {'-'*10} {'-'*12}"
    )
    for name, m in all_metrics:
        t1c = m.label_change["t1_coord"]
        cv = m.label_change["coord_cv"]
        print(
            f"  {name:<10} "
            f"{t1c.changed:>8} {t1c.fix_rate*100:>9.1f}% {t1c.break_rate*100:>11.1f}%  "
            f"{cv.changed:>8} {cv.fix_rate*100:>9.1f}% {cv.break_rate*100:>11.1f}%"
        )


# ── Visualisation: EM & BA by stage (Figure 4) ───────────────────────────


def draw_em_ba_chart(all_metrics: list[tuple[str, Metrics]], out: Path):
    """Draw EM & BA across pipeline stages (Figure 4)."""
    nm = len(all_metrics)
    fig, axes = plt.subplots(1, nm, figsize=(5 * nm + 1, 4), squeeze=False)
    fig.suptitle("EM & BA Across Pipeline Stages",
                 fontsize=14, fontweight="bold", color="#2c3e50")

    for col, (model, sm) in enumerate(all_metrics):
        ax = axes[0, col]
        x = np.arange(3)
        w = 0.3
        b_em = ax.bar(x - w / 2, [e * 100 for e in sm.em], w, label="EM", color="#2980b9", alpha=0.85)
        b_ba = ax.bar(x + w / 2, [b * 100 for b in sm.ba], w, label="BA", color="#27ae60", alpha=0.85)
        for bars in [b_em, b_ba]:
            for bar in bars:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                         f"{bar.get_height():.1f}", ha="center", va="bottom", fontsize=7)
        for i in range(1, 3):
            de = (sm.em[i] - sm.em[i - 1]) * 100
            db = (sm.ba[i] - sm.ba[i - 1]) * 100
            ax.annotate(f"{de:+.1f}", xy=(i - w / 2, sm.em[i] * 100),
                         fontsize=6.5, color="#27ae60" if de > 0 else "#c0392b",
                         ha="center", va="bottom", xytext=(0, 12), textcoords="offset points")
            ax.annotate(f"{db:+.1f}", xy=(i + w / 2, sm.ba[i] * 100),
                         fontsize=6.5, color="#27ae60" if db > 0 else "#c0392b",
                         ha="center", va="bottom", xytext=(0, 12), textcoords="offset points")
        ax.set_xticks(x)
        ax.set_xticklabels(STAGE_NAMES, fontsize=8)
        ax.set_ylim(0, 105)
        ax.set_ylabel("% rate", fontsize=8)
        ax.set_title(f"{model}", fontsize=10, fontweight="bold")
        ax.legend(fontsize=7, loc="lower right")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    plt.tight_layout(rect=[0, 0, 1, 0.92])
    fig.savefig(out, dpi=200, bbox_inches="tight", facecolor="white")
    print(f"Saved → {out}")


# ── Main ──────────────────────────────────────────────────────────────────


@app.command()
def main(
    output_dir: Path = typer.Option(PROJECT_ROOT / "data" / "results", "--output-dir", help="Directory for all outputs"),
    min_instances: int = typer.Option(5, "--min-instances"),
    data: Path = typer.Option(DATA_PATH, "--data"),
    models: str = typer.Option(",".join(MODEL_DIRS.keys()), "--models"),
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    human = load_human_labels(data)
    print(f"Loaded {len(human)} human-labeled instances")

    all_metrics: list[tuple[str, Metrics]] = []
    all_instances: dict[str, list[Instance]] = {}

    for model_key in [m.strip() for m in models.split(",")]:
        if model_key not in MODEL_DIRS:
            print(f"  [{model_key}] Unknown, skipping")
            continue
        instances = load_model_data(model_key, human)
        if len(instances) < min_instances:
            print(f"  [{model_key}] {len(instances)} instances (< {min_instances}), skipping")
            continue
        sm = compute_metrics(instances)
        all_metrics.append((model_key, sm))
        all_instances[model_key] = instances
        print_model_report(model_key, sm)
        write_csv(instances, model_key, output_dir)

    if not all_metrics:
        print("No models with enough data.")
        raise typer.Exit(1)

    print_cross_model_table(all_metrics)

    # Figure 4: EM & BA across pipeline stages
    draw_em_ba_chart(all_metrics, output_dir / "em_ba_by_stage.png")


if __name__ == "__main__":
    app()
