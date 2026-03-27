"""Evaluate misalignment detection across all methods and models.

Methods:
  - zero-shot, few-shot, cot: LLM baselines (output in misalign_output_data/<method>/)
  - mini-swe-agent: single-agent system (output in misalign_output_data/mini-swe-agent/)
  - paichecker: PAIChecker multi-agent system (output in misalign_output_data/checker/)

Usage:
  python evaluate.py                          # evaluate all available methods/models
  python evaluate.py --methods cot,paichecker --models gpt,claude
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import typer

app = typer.Typer()

PROJECT_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = PROJECT_ROOT / "data" / "swe_gym_all.csv"
DATA_DIR = PROJECT_ROOT / "data" / "outputs"

ALL_LABELS = ["SC", "FP", "DP", "IS", "UL", "Others", "No Misalignment"]
METHODS = ["zero-shot", "few-shot", "cot", "mini-swe-agent", "paichecker"]
MODELS = ["gpt", "claude", "gemini", "qwen"]

MODEL_PRICING: dict[str, tuple[float, float, float]] = {
    "gpt": (1.75, 0.175, 14.0),
    "claude": (3.0, 0.3, 15.0),
    "gemini": (2.0, 0.2, 12.0),
    "qwen": (0.4, 0.04, 2.4),
}


# ── Helpers ────────────────────────────────────────────────────────────────

def strip_suffix(label: str) -> str:
    return re.sub(r"-\d+$", "", label.strip())


def parse_human_label(raw: str | None) -> set[str]:
    if not raw or str(raw).strip() == "":
        return {"No Misalignment"}
    return {strip_suffix(p) for p in str(raw).split(",")}


def parse_model_labels(classifications: list[dict]) -> set[str]:
    return {c["label"] for c in classifications} if classifications else set()


def format_label_set(labels: set[str]) -> str:
    order = {l: i for i, l in enumerate(ALL_LABELS)}
    return ", ".join(sorted(labels, key=lambda x: order.get(x, 99)))


def _resolve_model_key(model_name: str) -> str:
    name = model_name.lower()
    for key in ("claude", "gemini", "qwen", "gpt"):
        if key in name:
            return key
    return "gpt"


def _estimate_cost_from_usage(
    *,
    prompt_tokens: int,
    cached_input_tokens: int,
    completion_tokens: int,
    model_name: str,
) -> tuple[float, int]:
    input_price, cached_input_price, output_price = MODEL_PRICING[_resolve_model_key(model_name)]
    non_cached_input_tokens = max(prompt_tokens - cached_input_tokens, 0)
    cost = (
        (non_cached_input_tokens / 1_000_000) * input_price
        + (cached_input_tokens / 1_000_000) * cached_input_price
        + (completion_tokens / 1_000_000) * output_price
    )
    return cost, non_cached_input_tokens


# ── Data loading ───────────────────────────────────────────────────────────

def load_human_labels(path: Path = XLSX_PATH) -> dict[str, set[str]]:
    if path.suffix.lower() == ".csv":
        result = {}
        with path.open(encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                if row.get("instance_id"):
                    result[row["instance_id"]] = parse_human_label(row.get("human_label"))

        return result
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    id_col, human_col = headers.index("instance_id"), headers.index("human_label")
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[id_col]:
            result[row[id_col]] = parse_human_label(row[human_col])
    wb.close()
    return result


def _find_output_file(method: str, model: str) -> Path | None:
    if method == "mini-swe-agent":
        candidates = [DATA_DIR / "mini-swe-agent" / f"misalign_outputs_{model}_all.jsonl"]
    elif method == "paichecker":
        candidates = [DATA_DIR / "checker" / f"misalign_outputs_{model}_all.jsonl"]
    else:
        candidates = [DATA_DIR / method / f"misalign_outputs_{model}_all.jsonl"]
    for p in candidates:
        if p.exists():
            return p
    return None


def load_model_outputs(path: Path) -> dict[str, dict]:
    all_records: dict[str, list[dict]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        if rec.get("status") in ("skipped", "attempt_failed"):
            all_records.setdefault(rec["instance_id"], []).append(rec)
            continue
        all_records.setdefault(rec["instance_id"], []).append(rec)
    results = {}
    for iid, records in all_records.items():
        successful = [r for r in records if r.get("status") not in ("skipped", "attempt_failed")]
        if successful:
            results[iid] = successful[0]
    return results


def load_attempt_details(path: Path) -> dict[str, list[dict]]:
    """Load all attempt records grouped by instance_id."""
    all_records: dict[str, list[dict]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        if rec.get("status") == "skipped":
            continue
        all_records.setdefault(rec["instance_id"], []).append(rec)
    return all_records


def print_attempt_stats(attempt_details: dict[str, list[dict]], method: str, model: str) -> None:
    has_attempt_info = any(
        r.get("attempt") is not None
        for records in attempt_details.values()
        for r in records
    )
    if not has_attempt_info:
        return
    first_try = 0
    retry_success = 0
    all_failed = 0
    attempt_distribution: dict[int, int] = {}
    per_instance_details: list[str] = []
    for iid, records in sorted(attempt_details.items()):
        failed = [r for r in records if r.get("status") == "attempt_failed"]
        successful = [r for r in records if r.get("status") not in ("skipped", "attempt_failed")]
        if successful:
            attempt_num = successful[0].get("attempt", 1)
            attempt_distribution[attempt_num] = attempt_distribution.get(attempt_num, 0) + 1
            if attempt_num == 1:
                first_try += 1
            else:
                retry_success += 1
                classifications = successful[0].get("classifications", [])
                labels = ", ".join(c["label"] for c in classifications) if classifications else "(none)"
                detail = f"    {iid}: OK on attempt {attempt_num}, labels=[{labels}]"
                for f in failed:
                    detail += f"\n      attempt {f.get('attempt', '?')}: FAILED - {f.get('error', 'unknown')[:80]}"
                per_instance_details.append(detail)
        elif failed:
            all_failed += 1
            detail = f"    {iid}: ALL FAILED ({len(failed)} attempts)"
            for f in failed:
                detail += f"\n      attempt {f.get('attempt', '?')}: {f.get('error', 'unknown')[:80]}"
            per_instance_details.append(detail)
    total = first_try + retry_success + all_failed
    print(f"\n  Attempt Statistics ({method}/{model}):")
    print(f"    1st attempt success: {first_try}/{total}")
    print(f"    Required retry:      {retry_success}/{total}")
    print(f"    All attempts failed: {all_failed}/{total}")
    for attempt_num in sorted(attempt_distribution):
        print(f"    Succeeded on attempt {attempt_num}: {attempt_distribution[attempt_num]}")
    if per_instance_details:
        print(f"\n  Per-instance attempt details ({method}/{model}):")
        for d in per_instance_details:
            print(d)


# ── Metrics ────────────────────────────────────────────────────────────────

@dataclass
class EvalResult:
    method: str
    model: str
    n: int = 0
    exact_match: int = 0
    per_label: dict[str, dict[str, int | float]] = field(default_factory=dict)
    instance_correct: dict[str, bool] = field(default_factory=dict)
    instance_labels: dict[str, str] = field(default_factory=dict)
    binary_accuracy: float = 0.0
    binary_precision: float = 0.0
    binary_recall: float = 0.0
    binary_f1: float = 0.0
    estimated_cost_usd: float = 0.0


def compute_metrics(human: dict[str, set[str]], model_outputs: dict[str, dict], method: str, model: str) -> EvalResult:
    result = EvalResult(method=method, model=model)
    paired = []
    for rec in model_outputs.values():
        usage = rec.get("token_usage", {})
        prompt_tokens = int(usage.get("prompt_tokens", 0) or 0)
        cached_input_tokens = int(usage.get("cached_input_tokens", 0) or 0)
        completion_tokens = int(usage.get("completion_tokens", 0) or 0)
        cost, _ = _estimate_cost_from_usage(
            prompt_tokens=prompt_tokens,
            cached_input_tokens=cached_input_tokens,
            completion_tokens=completion_tokens,
            model_name=model,
        )
        result.estimated_cost_usd += cost
    for iid, h_labels in human.items():
        classifications = (model_outputs.get(iid) or {}).get("classifications") or []
        has_prediction = bool(classifications)
        m_labels = parse_model_labels(classifications) if has_prediction else set()
        paired.append({"instance_id": iid, "human": h_labels, "model": m_labels, "has_prediction": has_prediction})

    result.n = len(human)
    result.exact_match = sum(1 for p in paired if p["human"] == p["model"])
    tp = tn = fp = fn = 0
    for p in paired:
        human_positive = p["human"] != {"No Misalignment"}
        model_positive = p["model"] != {"No Misalignment"}
        if human_positive and model_positive:
            tp += 1
        elif (not human_positive) and (not model_positive):
            tn += 1
        elif (not human_positive) and model_positive:
            fp += 1
        else:
            fn += 1
    result.binary_accuracy = (tp + tn) / result.n if result.n else 0.0
    result.binary_precision = tp / (tp + fp) if (tp + fp) else 0.0
    result.binary_recall = tp / (tp + fn) if (tp + fn) else 0.0
    result.binary_f1 = (
        2 * result.binary_precision * result.binary_recall / (result.binary_precision + result.binary_recall)
        if (result.binary_precision + result.binary_recall)
        else 0.0
    )
    for p in paired:
        result.instance_correct[p["instance_id"]] = p["human"] == p["model"]
        result.instance_labels[p["instance_id"]] = format_label_set(p["model"]) if p["has_prediction"] else ""

    for label in ALL_LABELS:
        tp = sum(1 for p in paired if label in p["human"] and label in p["model"])
        tn = sum(1 for p in paired if label not in p["human"] and label not in p["model"])
        fp = sum(1 for p in paired if label not in p["human"] and label in p["model"])
        fn = sum(1 for p in paired if label in p["human"] and label not in p["model"])
        total = tp + tn + fp + fn
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) else 0.0
        result.per_label[label] = {
            "tp": tp, "tn": tn, "fp": fp, "fn": fn,
            "accuracy": (tp + tn) / total if total else 0.0,
            "precision": precision, "recall": recall, "f1": f1,
        }
    return result


# ── Pretty print ───────────────────────────────────────────────────────────

def _active_labels(r: EvalResult) -> list[str]:
    return [
        l for l in ALL_LABELS
        if r.per_label.get(l, {}).get("tp", 0) + r.per_label.get(l, {}).get("fn", 0) + r.per_label.get(l, {}).get("fp", 0) > 0
        or l == "No Misalignment"
    ]


def _macro(r: EvalResult, key: str) -> float:
    labels = _active_labels(r)
    return sum(r.per_label[l][key] for l in labels) / len(labels) if labels else 0.0


def print_single_result(r: EvalResult) -> None:
    print(f"\n{'='*70}")
    print(f"  {r.method} / {r.model}  (n={r.n})")
    print(f"{'='*70}")
    print(f"  Exact Match: {r.exact_match / r.n:.4f}  ({r.exact_match}/{r.n})" if r.n else "  No data")
    print(f"  Binary Accuracy: {r.binary_accuracy:.4f}" if r.n else "")
    print(
        f"  Binary Precision/Recall/F1: {r.binary_precision:.4f}/{r.binary_recall:.4f}/{r.binary_f1:.4f}"
        if r.n
        else ""
    )
    print(f"\n  {'Label':<20} {'Acc':>6} {'Prec':>6} {'Rec':>6} {'F1':>6}")
    for label in ALL_LABELS:
        s = r.per_label.get(label, {})
        if not s:
            continue
        if s["tp"] + s["fn"] + s["fp"] > 0 or label == "No Misalignment":
            print(f"  {label:<20} {s['accuracy']:>6.3f} {s['precision']:>6.3f} {s['recall']:>6.3f} {s['f1']:>6.3f}")
    print(f"  {'Macro Accuracy':<20} {_macro(r, 'accuracy'):>6.3f}")
    print(f"  {'Macro Precision':<20} {_macro(r, 'precision'):>6.3f}")
    print(f"  {'Macro Recall':<20} {_macro(r, 'recall'):>6.3f}")
    print(f"  {'Macro F1':<20} {_macro(r, 'f1'):>6.3f}")
    print(f"  Estimated Cost (USD): {r.estimated_cost_usd:.6f}")


def print_comparison_table(results: list[EvalResult]) -> None:
    print(f"\n{'='*130}")
    print("  COMPARISON TABLE")
    print(f"{'='*130}")
    print(f"  {'Method':<20} {'Model':<10} {'N':>4} {'EM':>6} {'EM%':>7} {'BA':>7} {'BP':>7} {'BR':>7} {'BF1':>7} {'MacroAcc':>9} {'MacroPrec':>10} {'MacroRec':>9} {'MacroF1':>8} {'CostUSD':>10}")
    print(f"  {'-'*20} {'-'*10} {'-'*4} {'-'*6} {'-'*7} {'-'*7} {'-'*7} {'-'*7} {'-'*7} {'-'*9} {'-'*10} {'-'*9} {'-'*8} {'-'*10}")
    for r in results:
        print(
            f"  {r.method:<20} {r.model:<10} {r.n:>4} {r.exact_match:>6}"
            f" {r.exact_match / r.n if r.n else 0:>7.4f}"
            f" {r.binary_accuracy:>7.4f}"
            f" {r.binary_precision:>7.4f} {r.binary_recall:>7.4f} {r.binary_f1:>7.4f}"
            f" {_macro(r, 'accuracy'):>9.4f} {_macro(r, 'precision'):>10.4f}"
            f" {_macro(r, 'recall'):>9.4f} {_macro(r, 'f1'):>8.4f}"
            f" {r.estimated_cost_usd:>10.6f}"
        )


# ── Excel output ───────────────────────────────────────────────────────────

def write_detail_excel(human: dict[str, set[str]], results: list[EvalResult], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Instance Details"

    combos = [(r.method, r.model) for r in results if r.n > 0]
    all_ids = sorted(human.keys())

    combo_headers = []
    for method, model in combos:
        prefix = f"{method}_{model}"
        combo_headers.extend([f"{prefix}_correct", f"{prefix}_labels"])
    headers = ["instance_id", "human_label"] + combo_headers + ["correct_count", "total_methods"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    lookup_correct = {(r.method, r.model): r.instance_correct for r in results if r.n > 0}
    lookup_labels = {(r.method, r.model): r.instance_labels for r in results if r.n > 0}

    for row, iid in enumerate(all_ids, 2):
        ws.cell(row=row, column=1, value=iid)
        ws.cell(row=row, column=2, value=format_label_set(human[iid]))
        correct = total = 0
        for ci, (method, model) in enumerate(combos):
            ic = lookup_correct.get((method, model), {})
            il = lookup_labels.get((method, model), {})
            base_col = 3 + ci * 2
            if iid in ic:
                is_correct = ic[iid]
                ws.cell(row=row, column=base_col, value="yes" if is_correct else "no")
                ws.cell(row=row, column=base_col + 1, value=il.get(iid, ""))
                correct += int(is_correct)
                total += 1
            else:
                ws.cell(row=row, column=base_col, value="N/A")
                ws.cell(row=row, column=base_col + 1, value="")
        ws.cell(row=row, column=3 + len(combos) * 2, value=correct)
        ws.cell(row=row, column=4 + len(combos) * 2, value=total)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    for ci, h in enumerate(
        [
            "method",
            "model",
            "n",
            "exact_match",
            "exact_match_rate",
            "binary_accuracy",
            "binary_precision",
            "binary_recall",
            "binary_f1",
            "macro_accuracy",
            "macro_precision",
            "macro_recall",
            "macro_f1",
            "estimated_cost_usd",
        ],
        1,
    ):
        ws2.cell(row=1, column=ci, value=h)
    row = 2
    for r in results:
        if r.n == 0:
            continue
        ws2.cell(row=row, column=1, value=r.method)
        ws2.cell(row=row, column=2, value=r.model)
        ws2.cell(row=row, column=3, value=r.n)
        ws2.cell(row=row, column=4, value=r.exact_match)
        ws2.cell(row=row, column=5, value=r.exact_match / r.n if r.n else 0.0)
        ws2.cell(row=row, column=6, value=r.binary_accuracy)
        ws2.cell(row=row, column=7, value=r.binary_precision)
        ws2.cell(row=row, column=8, value=r.binary_recall)
        ws2.cell(row=row, column=9, value=r.binary_f1)
        ws2.cell(row=row, column=10, value=_macro(r, "accuracy"))
        ws2.cell(row=row, column=11, value=_macro(r, "precision"))
        ws2.cell(row=row, column=12, value=_macro(r, "recall"))
        ws2.cell(row=row, column=13, value=_macro(r, "f1"))
        ws2.cell(row=row, column=14, value=r.estimated_cost_usd)
        row += 1

    wb.save(output_path)
    print(f"\nExcel saved → {output_path}")


# ── Main ───────────────────────────────────────────────────────────────────

@app.command()
def main(
    output_xlsx: Path = typer.Option(PROJECT_ROOT / "data" / "results" / "evaluation_results.xlsx", "--output", help="Output Excel path"),
    methods: str = typer.Option(",".join(METHODS), "--methods", help="Comma-separated methods"),
    models: str = typer.Option(",".join(MODELS), "--models", help="Comma-separated models"),
    human_labels: Path = typer.Option(XLSX_PATH, "--human-labels", help="Human labels file (.xlsx or .csv)"),
) -> None:
    method_list = [m.strip() for m in methods.split(",")]
    model_list = [m.strip() for m in models.split(",")]

    human = load_human_labels(human_labels)
    print(f"Loaded {len(human)} human-labeled instances")

    all_results: list[EvalResult] = []
    for method in method_list:
        for model in model_list:
            path = _find_output_file(method, model)
            if path is None:
                print(f"  [{method}/{model}] No output file found, skipping")
                continue
            outputs = load_model_outputs(path)
            if not outputs:
                print(f"  [{method}/{model}] Empty output file, evaluating against all instances with empty predictions")
            r = compute_metrics(human, outputs, method, model)
            all_results.append(r)
            print_single_result(r)
            # attempt_details = load_attempt_details(path)
            # print_attempt_stats(attempt_details, method, model)

    if all_results:
        print_comparison_table(all_results)
        write_detail_excel(human, all_results, output_xlsx)


if __name__ == "__main__":
    app()
